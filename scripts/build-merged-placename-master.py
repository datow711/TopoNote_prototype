#!/usr/bin/env python3
"""Create a local Places master enriched only by conservative one-to-one matches."""

from __future__ import annotations

import argparse
import html
import json
import math
import re
import shutil
import unicodedata
from collections import Counter, defaultdict
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


TYPE_CODE_TO_LABEL = {
    "1": "行政區域",
    "2": "聚落",
    "3": "自然地理實體",
    "4": "具有地標意義公共設施",
    "5": "街道",
    "6": "其他",
}
TYPE_LABEL_TO_CODE = {label: code for code, label in TYPE_CODE_TO_LABEL.items()}

ADDED_HEADERS = [
    "LocationDescription",
    "HistoryDescription",
    "StandardPlaceCode",
    "DataSource",
    "SourceUID",
    "SourcePlaceId",
    "MergeMatchMethod",
    "CoordinateUpdateStatus",
    "MergeBatch",
]


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return unicodedata.normalize("NFKC", str(value)).strip()


def clean_source_text(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value))
    text = text.replace("_x000D_", "\n").replace("_x000A_", "\n")
    text = html.unescape(text).replace("\r\n", "\n").replace("\r", "\n")
    text = "\n".join(line.rstrip() for line in text.split("\n"))
    return text.strip()


def parse_number(value: Any) -> float | None:
    if value is None or normalize_text(value) == "":
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def normalize_coordinates(latitude: Any, longitude: Any) -> tuple[float | None, float | None, str]:
    lat = parse_number(latitude)
    lon = parse_number(longitude)
    if lat is None or lon is None or lat == 0 or lon == 0:
        return None, None, "NEW_INVALID_PRESERVED_ORIGINAL"
    if abs(lat) > 90 and abs(lon) <= 90:
        lat, lon = lon, lat
        if abs(lat) <= 90 and abs(lon) <= 180:
            return lat, lon, "NEW_SWAPPED_CORRECTED"
    if abs(lat) <= 90 and abs(lon) <= 180:
        return lat, lon, "NEW_VALID_UPDATED"
    return None, None, "NEW_INVALID_PRESERVED_ORIGINAL"


def load_new_inventory(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["總表"]
    iterator = sheet.iter_rows(values_only=True)
    headers = [normalize_text(value) for value in next(iterator)]
    rows = []
    for source_row, values in enumerate(iterator, start=2):
        row = {headers[index]: values[index] if index < len(values) else None for index in range(len(headers))}
        row["__source_row__"] = source_row
        rows.append(row)
    workbook.close()
    return headers, rows


def copy_cell_style(source, target) -> None:
    if source.has_style:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def build(args: argparse.Namespace) -> dict[str, Any]:
    _, new_rows = load_new_inventory(args.new_inventory)
    new_index: dict[tuple[str, str, str], list[tuple[int, dict[str, Any]]]] = defaultdict(list)
    for source_index, row in enumerate(new_rows):
        key = (
            normalize_text(row.get("*所屬縣市")),
            normalize_text(row.get("*所屬鄉鎮")),
            normalize_text(row.get("*地名名稱(中文)")),
        )
        if key[0] and key[2]:
            new_index[key].append((source_index, row))

    args.output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(args.base_master, args.output)
    workbook = load_workbook(args.output, data_only=False)
    sheet = workbook["比對結果_first_million"]
    headers = [normalize_text(cell.value) for cell in sheet[1]]
    header_index = {header: index + 1 for index, header in enumerate(headers)}

    provisional = []
    source_to_base_rows: dict[int, list[int]] = defaultdict(list)
    raw_status = Counter()
    for row_number in range(2, sheet.max_row + 1):
        key = (
            normalize_text(sheet.cell(row_number, header_index["County"]).value),
            normalize_text(sheet.cell(row_number, header_index["Town"]).value),
            normalize_text(sheet.cell(row_number, header_index["PlaceName"]).value),
        )
        candidates = new_index.get(key, []) if key[0] and key[2] else []
        selected = None
        status = "NO_MATCH"
        if len(candidates) == 1:
            selected = candidates[0]
            status = "DIRECT_UNIQUE"
        elif len(candidates) > 1:
            expected_code = TYPE_LABEL_TO_CODE.get(
                normalize_text(sheet.cell(row_number, header_index["Type"]).value), ""
            )
            filtered = [
                candidate
                for candidate in candidates
                if normalize_text(candidate[1].get("*地名類別(請填代號)")) == expected_code
            ]
            if expected_code and len(filtered) == 1:
                selected = filtered[0]
                status = "CATEGORY_RESOLVED"
            elif expected_code and not filtered:
                status = "CATEGORY_MISMATCH"
            else:
                status = "AMBIGUOUS"
        provisional.append((row_number, status, selected))
        raw_status[status] += 1
        if selected is not None:
            source_to_base_rows[selected[0]].append(row_number)

    final_status = Counter()
    safe_matches: dict[int, tuple[str, dict[str, Any]]] = {}
    for row_number, status, selected in provisional:
        if selected is None:
            final_status[status] += 1
        elif len(source_to_base_rows[selected[0]]) == 1:
            final_status[status] += 1
            safe_matches[row_number] = (status, selected[1])
        else:
            final_status["SOURCE_COLLISION"] += 1

    first_added_column = sheet.max_column + 1
    exemplar = sheet.cell(1, sheet.max_column)
    for offset, header in enumerate(ADDED_HEADERS):
        cell = sheet.cell(1, first_added_column + offset, header)
        copy_cell_style(exemplar, cell)
        cell.alignment = copy(cell.alignment)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    added_index = {header: first_added_column + index for index, header in enumerate(ADDED_HEADERS)}
    coordinate_status = Counter()
    uuid_before = [sheet.cell(row, header_index["序號"]).value for row in range(2, sheet.max_row + 1)]
    for row_number, (match_method, source) in safe_matches.items():
        sheet.cell(row_number, header_index["Village"], clean_source_text(source.get("所屬村里")) or None)
        sheet.cell(row_number, added_index["LocationDescription"], clean_source_text(source.get("相關位置與面積描述")) or None)
        sheet.cell(row_number, added_index["HistoryDescription"], clean_source_text(source.get("地名沿革與文獻歷史簡述")) or None)
        sheet.cell(row_number, added_index["StandardPlaceCode"], clean_source_text(source.get("標準地名代碼")) or None)
        sheet.cell(row_number, added_index["DataSource"], clean_source_text(source.get("資料來源")) or None)
        sheet.cell(row_number, added_index["SourceUID"], normalize_text(source.get("地名UID")) or None)
        sheet.cell(row_number, added_index["SourcePlaceId"], normalize_text(source.get("*PlaceId")) or None)
        sheet.cell(row_number, added_index["MergeMatchMethod"], match_method)
        sheet.cell(row_number, added_index["MergeBatch"], args.batch_name)

        latitude, longitude, coord_status = normalize_coordinates(source.get("*緯度"), source.get("*經度"))
        if latitude is not None and longitude is not None:
            sheet.cell(row_number, header_index["Longitude"], longitude)
            sheet.cell(row_number, header_index["Latitude"], latitude)
        sheet.cell(row_number, added_index["CoordinateUpdateStatus"], coord_status)
        coordinate_status[coord_status] += 1

    uuid_after = [sheet.cell(row, header_index["序號"]).value for row in range(2, sheet.max_row + 1)]
    if uuid_before != uuid_after:
        raise RuntimeError("UUID column changed unexpectedly")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
    widths = {
        "LocationDescription": 42,
        "HistoryDescription": 60,
        "StandardPlaceCode": 18,
        "DataSource": 16,
        "SourceUID": 38,
        "SourcePlaceId": 20,
        "MergeMatchMethod": 22,
        "CoordinateUpdateStatus": 32,
        "MergeBatch": 26,
    }
    for header, width in widths.items():
        sheet.column_dimensions[get_column_letter(added_index[header])].width = width

    summary_name = "合併摘要_1150112"
    if summary_name in workbook.sheetnames:
        del workbook[summary_name]
    summary_sheet = workbook.create_sheet(summary_name, 1)
    summary_sheet.append(["項目", "筆數／內容"])
    summary_rows = [
        ("原 Places 筆數", sheet.max_row - 1),
        ("新清冊筆數", len(new_rows)),
        ("安全合併總數", len(safe_matches)),
        ("直接唯一候選", final_status["DIRECT_UNIQUE"]),
        ("同名多候選、類別唯一", final_status["CATEGORY_RESOLVED"]),
        ("跳過：仍有多候選", final_status["AMBIGUOUS"]),
        ("跳過：類別不相符", final_status["CATEGORY_MISMATCH"]),
        ("跳過：找不到候選", final_status["NO_MATCH"]),
        ("跳過：同一來源指向多筆原資料", final_status["SOURCE_COLLISION"]),
        ("座標：有效且已更新", coordinate_status["NEW_VALID_UPDATED"]),
        ("座標：經緯度反寫已校正", coordinate_status["NEW_SWAPPED_CORRECTED"]),
        ("座標：新值無效、保留原值", coordinate_status["NEW_INVALID_PRESERVED_ORIGINAL"]),
        ("比對鍵", "County + Town + PlaceName；多候選時再用 Type"),
        ("UUID", "全部保留原 Places 序號，未改動"),
        ("批次", args.batch_name),
    ]
    for row in summary_rows:
        summary_sheet.append(row)
    summary_sheet.freeze_panes = "A2"
    summary_sheet.auto_filter.ref = f"A1:B{summary_sheet.max_row}"
    summary_sheet.column_dimensions["A"].width = 34
    summary_sheet.column_dimensions["B"].width = 58
    summary_sheet["A1"].fill = PatternFill("solid", fgColor="D9EAF7")
    summary_sheet["B1"].fill = PatternFill("solid", fgColor="D9EAF7")
    summary_sheet["A1"].font = Font(bold=True)
    summary_sheet["B1"].font = Font(bold=True)
    for row in summary_sheet.iter_rows(min_row=1, max_row=summary_sheet.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    workbook.save(args.output)
    workbook.close()

    summary = {
        "output": str(args.output),
        "base_rows": len(uuid_before),
        "new_rows": len(new_rows),
        "raw_status": dict(raw_status),
        "final_status": dict(final_status),
        "safe_merge_rows": len(safe_matches),
        "coordinate_status": dict(coordinate_status),
        "uuid_preserved": True,
        "added_headers": ADDED_HEADERS,
    }
    args.summary_json.parent.mkdir(parents=True, exist_ok=True)
    args.summary_json.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--base-master", type=Path, required=True)
    parser.add_argument("--new-inventory", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--summary-json", type=Path, required=True)
    parser.add_argument("--batch-name", default="115.1.12_地名後臺清冊_合併")
    args = parser.parse_args()
    print(json.dumps(build(args), ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
