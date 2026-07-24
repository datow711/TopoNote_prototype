#!/usr/bin/env python3
"""Match rows between the placename master workbook and backend inventory."""

from __future__ import annotations

import argparse
import html
import json
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


TYPE_CODE_TO_LABEL = {
    "1": "行政區域",
    "2": "聚落",
    "3": "自然地理實體",
    "4": "具有地標意義公共設施",
    "5": "街道",
}

MASTER_SHEETS = [
    {
        "sheet": "Places",
        "id_header": "序號",
        "name_header": "PlaceName",
        "output_prefix": "Places",
    },
    {
        "sheet": "地名總表",
        "id_header": "UUID",
        "name_header": "地名",
        "output_prefix": "Master",
    },
]

PAIR_HEADERS = [
    "MatchStatus",
    "MatchKey",
    "MasterSheet",
    "MasterRow",
    "MasterId",
    "MasterType",
    "MasterPlaceName",
    "MasterCounty",
    "MasterTown",
    "InventoryRow",
    "InventoryUID",
    "InventoryPlaceId",
    "InventoryCategoryCode",
    "InventoryMappedType",
    "InventoryPlaceName",
    "InventoryCounty",
    "InventoryTown",
    "InventoryRowsForKey",
    "MasterRowsForKey",
]

UNMATCHED_HEADERS = [
    "InventoryRow",
    "InventoryUID",
    "InventoryPlaceId",
    "InventoryCategoryCode",
    "InventoryMappedType",
    "InventoryPlaceName",
    "InventoryCounty",
    "InventoryTown",
    "Reason",
]


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", html.unescape(str(value)))
    return re.sub(r"\s+", " ", text).strip()


def read_header_index(sheet) -> dict[str, int]:
    return {normalize_text(cell.value): index for index, cell in enumerate(sheet[1])}


def load_inventory(path: Path) -> dict[tuple[str, str, str, str], list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["總表"]
    index = read_header_index(sheet)
    required = [
        "地名UID",
        "*PlaceId",
        "*地名名稱(中文)",
        "*所屬縣市",
        "*所屬鄉鎮",
        "*地名類別(請填代號)",
    ]
    missing = [header for header in required if header not in index]
    if missing:
        raise KeyError(f"Inventory sheet missing headers: {missing}")

    rows_by_key: dict[tuple[str, str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        code = normalize_text(values[index["*地名類別(請填代號)"]])
        mapped_type = TYPE_CODE_TO_LABEL.get(code, "")
        name = normalize_text(values[index["*地名名稱(中文)"]])
        county = normalize_text(values[index["*所屬縣市"]])
        town = normalize_text(values[index["*所屬鄉鎮"]])
        item = {
            "row": row_number,
            "uid": normalize_text(values[index["地名UID"]]),
            "place_id": normalize_text(values[index["*PlaceId"]]),
            "category_code": code,
            "mapped_type": mapped_type,
            "name": name,
            "county": county,
            "town": town,
        }
        key = (mapped_type, name, county, town)
        rows_by_key[key].append(item)
    workbook.close()
    return rows_by_key


def load_master(path: Path, sheet_config: dict[str, str]) -> dict[tuple[str, str, str, str], list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_config["sheet"]]
    index = read_header_index(sheet)
    required = ["Type", sheet_config["name_header"], "County", "Town", sheet_config["id_header"]]
    missing = [header for header in required if header not in index]
    if missing:
        raise KeyError(f"{sheet_config['sheet']} missing headers: {missing}")

    rows_by_key: dict[tuple[str, str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        type_label = normalize_text(values[index["Type"]])
        name = normalize_text(values[index[sheet_config["name_header"]]])
        county = normalize_text(values[index["County"]])
        town = normalize_text(values[index["Town"]])
        key = (type_label, name, county, town)
        if not all(key):
            continue
        rows_by_key[key].append(
            {
                "sheet": sheet_config["sheet"],
                "row": row_number,
                "id": normalize_text(values[index[sheet_config["id_header"]]]),
                "type": type_label,
                "name": name,
                "county": county,
                "town": town,
            }
        )
    workbook.close()
    return rows_by_key


def append_rows(sheet, headers: list[str], rows: list[list[Any]]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append(row)


def build_pair_rows(
    inventory_by_key: dict[tuple[str, str, str, str], list[dict[str, Any]]],
    master_by_key: dict[tuple[str, str, str, str], list[dict[str, Any]]],
) -> tuple[list[list[Any]], list[list[Any]], Counter]:
    pair_rows: list[list[Any]] = []
    unmatched_rows: list[list[Any]] = []
    counts: Counter = Counter()
    all_keys = set(inventory_by_key) | set(master_by_key)

    for key in sorted(all_keys):
        inventory_rows = inventory_by_key.get(key, [])
        master_rows = master_by_key.get(key, [])
        if not inventory_rows:
            continue

        mapped_type, name, county, town = key
        if not mapped_type:
            for inventory in inventory_rows:
                unmatched_rows.append(unmatched_inventory_row(inventory, "UNMAPPED_CATEGORY_CODE"))
                counts["inventory_unmapped_category"] += 1
            continue

        if not master_rows:
            for inventory in inventory_rows:
                unmatched_rows.append(unmatched_inventory_row(inventory, "NO_MASTER_KEY_MATCH"))
                counts["inventory_no_master_key_match"] += 1
            continue

        status = "UNIQUE_1_TO_1" if len(inventory_rows) == 1 and len(master_rows) == 1 else "AMBIGUOUS_SAME_KEY"
        counts[status] += 1
        counts["matched_keys"] += 1
        counts["matched_inventory_rows"] += len(inventory_rows)
        counts["matched_master_rows"] += len(master_rows)
        key_text = " | ".join([mapped_type, name, county, town])

        for master in master_rows:
            for inventory in inventory_rows:
                pair_rows.append(
                    [
                        status,
                        key_text,
                        master["sheet"],
                        master["row"],
                        master["id"],
                        master["type"],
                        master["name"],
                        master["county"],
                        master["town"],
                        inventory["row"],
                        inventory["uid"],
                        inventory["place_id"],
                        inventory["category_code"],
                        inventory["mapped_type"],
                        inventory["name"],
                        inventory["county"],
                        inventory["town"],
                        len(inventory_rows),
                        len(master_rows),
                    ]
                )
                counts["candidate_pairs"] += 1

    return pair_rows, unmatched_rows, counts


def unmatched_inventory_row(inventory: dict[str, Any], reason: str) -> list[Any]:
    return [
        inventory["row"],
        inventory["uid"],
        inventory["place_id"],
        inventory["category_code"],
        inventory["mapped_type"],
        inventory["name"],
        inventory["county"],
        inventory["town"],
        reason,
    ]


def write_report(args: argparse.Namespace) -> dict[str, Any]:
    inventory_by_key = load_inventory(args.inventory)
    workbook = Workbook(write_only=True)
    summary_rows = [
        ["Item", "Value"],
        ["Inventory workbook", str(args.inventory)],
        ["Master workbook", str(args.master)],
        ["Inventory sheet", "總表"],
        ["Comparison fields", "Type/category + PlaceName/name + County + Town"],
    ]
    type_mapping_rows = [["InventoryCategoryCode", "MappedMasterType"]]
    for code, label in TYPE_CODE_TO_LABEL.items():
        type_mapping_rows.append([code, label])
    type_mapping_rows.append(["6", "UNMAPPED_CODE"])

    summary: dict[str, Any] = {"sheets": {}, "type_code_to_label": {**TYPE_CODE_TO_LABEL, "6": "UNMAPPED_CODE"}}

    for config in MASTER_SHEETS:
        master_by_key = load_master(args.master, config)
        pair_rows, unmatched_rows, counts = build_pair_rows(inventory_by_key, master_by_key)
        unique_rows = [row for row in pair_rows if row[0] == "UNIQUE_1_TO_1"]
        ambiguous_rows = [row for row in pair_rows if row[0] == "AMBIGUOUS_SAME_KEY"]
        prefix = config["output_prefix"]

        append_rows(workbook.create_sheet(f"{prefix}_Unique"), PAIR_HEADERS, unique_rows)
        append_rows(workbook.create_sheet(f"{prefix}_Ambiguous"), PAIR_HEADERS, ambiguous_rows)
        append_rows(workbook.create_sheet(f"{prefix}_UnmatchedInventory"), UNMATCHED_HEADERS, unmatched_rows)

        counts["unique_candidate_pairs"] = len(unique_rows)
        counts["ambiguous_candidate_pairs"] = len(ambiguous_rows)
        counts["unmatched_inventory_rows"] = len(unmatched_rows)
        summary["sheets"][config["sheet"]] = dict(counts)

        summary_rows.extend(
            [
                [f"{config['sheet']} matched keys", counts["matched_keys"]],
                [f"{config['sheet']} unique 1-to-1 keys", counts["UNIQUE_1_TO_1"]],
                [f"{config['sheet']} ambiguous same-key groups", counts["AMBIGUOUS_SAME_KEY"]],
                [f"{config['sheet']} candidate pairs", counts["candidate_pairs"]],
                [f"{config['sheet']} unmatched inventory rows", len(unmatched_rows)],
            ]
        )

    append_rows(workbook.create_sheet("Summary"), ["Item", "Value"], summary_rows[1:])
    append_rows(workbook.create_sheet("TypeMapping"), ["InventoryCategoryCode", "MappedMasterType"], type_mapping_rows[1:])
    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]

    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)
    if args.summary_json:
        args.summary_json.parent.mkdir(parents=True, exist_ok=True)
        args.summary_json.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--master", type=Path, required=True)
    parser.add_argument("--inventory", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--summary-json", type=Path)
    args = parser.parse_args()
    print(json.dumps(write_report(args), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()


