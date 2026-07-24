#!/usr/bin/env python3
"""Build a conservative staging dataset for enriching the live place master.

The live master snapshot is a JSONL file whose rows are arrays in this order:
UUID, Type, PlaceName, County, Town, Longitude, Latitude, live row number.
The new inventory workbook must contain a worksheet named ``總表``.
"""

from __future__ import annotations

import argparse
import html
import json
import math
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


TYPE_CODE_TO_LABEL = {
    "1": "行政區域",
    "2": "聚落",
    "3": "自然地理實體",
    "4": "具有地標意義公共設施",
    "5": "街道",
    "6": "其他",
}

HEADERS = [
    "MatchStatus",
    "MatchMethod",
    "StagingUUID",
    "OriginalUUID",
    "NewSourceUID",
    "NewPlaceId",
    "NewCategoryCode",
    "Type",
    "PlaceName",
    "County",
    "Town",
    "Village",
    "LocationDescription",
    "HistoryDescription",
    "StandardPlaceCode",
    "DataSource",
    "OriginalLongitude",
    "OriginalLatitude",
    "ProposedLongitude",
    "ProposedLatitude",
    "CoordinateStatus",
    "CandidateCount",
    "NewRowsForTarget",
    "OriginalRow",
    "MatchNote",
    "ImportBatch",
]

COMPACT_HEADERS = [
    "MatchInfo",
    "UUID",
    "Village",
    "LocationDescription",
    "HistoryDescription",
    "StandardPlaceCode",
    "DataSource",
    "Coordinates",
]


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value))
    return re.sub(r"\s+", " ", text).strip()


def clean_source_text(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value))
    text = text.replace("_x000D_", "\n").replace("_x000A_", "\n")
    text = html.unescape(text).replace("\r\n", "\n").replace("\r", "\n")
    text = "\n".join(line.rstrip() for line in text.split("\n"))
    return text.strip()


def parse_number(value: Any) -> float | None:
    if value is None or normalize_text(value) == "":
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def normalize_coordinates(latitude: Any, longitude: Any) -> tuple[float | None, float | None, str]:
    lat = parse_number(latitude)
    lon = parse_number(longitude)
    if lat is None or lon is None or lat == 0 or lon == 0:
        return None, None, "NEW_INVALID_OR_ZERO"
    if abs(lat) > 90 and abs(lon) <= 90:
        lat, lon = lon, lat
        if abs(lat) <= 90 and abs(lon) <= 180:
            return lat, lon, "NEW_SWAPPED_CORRECTED"
    if abs(lat) <= 90 and abs(lon) <= 180:
        return lat, lon, "NEW_VALID_UPDATED"
    return None, None, "NEW_OUT_OF_RANGE"


def coordinates_are_close(
    new_coordinates: tuple[float | None, float | None, str],
    live_row: dict[str, Any],
    tolerance: float = 0.0002,
) -> bool:
    new_lat, new_lon, _ = new_coordinates
    live_lat = live_row["latitude"]
    live_lon = live_row["longitude"]
    if new_lat is None or new_lon is None or live_lat is None or live_lon is None:
        return False
    if new_lat == 0 or new_lon == 0 or live_lat == 0 or live_lon == 0:
        return False
    return abs(new_lat - live_lat) <= tolerance and abs(new_lon - live_lon) <= tolerance


def load_live_snapshot(path: Path) -> list[dict[str, Any]]:
    rows = []
    with path.open("r", encoding="utf-8") as source:
        for line in source:
            line = line.strip()
            if not line or line == "__END__":
                continue
            values = json.loads(line)
            rows.append(
                {
                    "uuid": normalize_text(values[0]),
                    "type": normalize_text(values[1]),
                    "name": normalize_text(values[2]),
                    "county": normalize_text(values[3]),
                    "town": normalize_text(values[4]),
                    "longitude": parse_number(values[5]),
                    "latitude": parse_number(values[6]),
                    "row": int(values[7]),
                }
            )
    return rows


def load_new_inventory(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["總表"]
    iterator = sheet.iter_rows(values_only=True)
    headers = [normalize_text(value) for value in next(iterator)]
    rows = []
    for source_row, values in enumerate(iterator, start=2):
        row = {headers[index]: values[index] if index < len(values) else None for index in range(len(headers))}
        row["__source_row__"] = source_row
        rows.append(row)
    workbook.close()
    return rows


def build_staging(live_rows: list[dict[str, Any]], new_rows: list[dict[str, Any]], batch_name: str):
    live_index: dict[tuple[str, str, str, str], list[dict[str, Any]]] = defaultdict(list)
    live_uuid_counts = Counter(row["uuid"] for row in live_rows if row["uuid"])
    for row in live_rows:
        key = (row["county"], row["town"], row["name"], row["type"])
        live_index[key].append(row)

    provisional = []
    target_to_source_indexes: dict[str, list[int]] = defaultdict(list)
    for source_index, row in enumerate(new_rows):
        source_uid = normalize_text(row.get("地名UID"))
        place_id = normalize_text(row.get("*PlaceId"))
        code = normalize_text(row.get("*地名類別(請填代號)"))
        type_label = TYPE_CODE_TO_LABEL.get(code, "")
        name = normalize_text(row.get("*地名名稱(中文)"))
        county = normalize_text(row.get("*所屬縣市"))
        town = normalize_text(row.get("*所屬鄉鎮"))
        key = (county, town, name, type_label)
        candidates = live_index.get(key, []) if all(key) else []
        item = {
            "source_index": source_index,
            "source_row": row["__source_row__"],
            "source_uid": source_uid,
            "place_id": place_id,
            "code": code,
            "type": type_label,
            "name": name,
            "county": county,
            "town": town,
            "village": clean_source_text(row.get("所屬村里")),
            "location": clean_source_text(row.get("相關位置與面積描述")),
            "history": clean_source_text(row.get("地名沿革與文獻歷史簡述")),
            "standard_code": clean_source_text(row.get("標準地名代碼")),
            "data_source": clean_source_text(row.get("資料來源")),
            "candidate_count": len(candidates),
            "candidates": candidates,
            "new_coordinates": normalize_coordinates(row.get("*緯度"), row.get("*經度")),
            "selected_target": None,
            "match_method": "",
        }
        if len(candidates) == 1:
            item["selected_target"] = candidates[0]
            item["match_method"] = "KEY_CATEGORY"
        elif len(candidates) > 1:
            close_candidates = [candidate for candidate in candidates if coordinates_are_close(item["new_coordinates"], candidate)]
            if len(close_candidates) == 1:
                item["selected_target"] = close_candidates[0]
                item["match_method"] = "KEY_CATEGORY_COORDINATE"
        provisional.append(item)
        if item["selected_target"] is not None:
            target_to_source_indexes[item["selected_target"]["uuid"]].append(source_index)

    target_winner_indexes: dict[str, int | None] = {}
    for target_uuid, source_indexes in target_to_source_indexes.items():
        if len(source_indexes) == 1:
            target_winner_indexes[target_uuid] = source_indexes[0]
            continue
        target = provisional[source_indexes[0]]["selected_target"]
        close_indexes = [
            source_index
            for source_index in source_indexes
            if coordinates_are_close(provisional[source_index]["new_coordinates"], target)
        ]
        target_winner_indexes[target_uuid] = close_indexes[0] if len(close_indexes) == 1 else None

    output_rows = []
    status_counts = Counter()
    coordinate_counts = Counter()
    overlong_cells = 0
    for item in provisional:
        candidates = item["candidates"]
        target = item["selected_target"]
        target_group_size = len(target_to_source_indexes[target["uuid"]]) if target is not None else 0
        winner_index = target_winner_indexes.get(target["uuid"]) if target is not None else None
        collision = target is not None and target_group_size > 1 and winner_index != item["source_index"]
        selected_unique = (
            target is not None
            and not collision
            and winner_index == item["source_index"]
            and live_uuid_counts[target["uuid"]] == 1
        )
        if selected_unique:
            match_status = "UNIQUE_MATCH"
            match_method = item["match_method"]
            if target_group_size > 1:
                match_method = "KEY_CATEGORY_COORDINATE_COLLISION_RESOLVED"
            staging_uuid = target["uuid"]
            original_uuid = target["uuid"]
            type_label = target["type"]
            name = target["name"]
            county = target["county"]
            town = target["town"]
            original_lon = target["longitude"]
            original_lat = target["latitude"]
            original_row = target["row"]
            note = "縣市、鄉鎮、地名與類別唯一對應；保留原 UUID。"
        elif collision:
            match_status = "AMBIGUOUS_TARGET_COLLISION"
            match_method = "KEY_CATEGORY_COLLISION_UNRESOLVED"
            staging_uuid = item["source_uid"]
            original_uuid = ""
            type_label, name, county, town = item["type"], item["name"], item["county"], item["town"]
            original_lon = original_lat = None
            original_row = ""
            note = "多筆新清冊資料指向同一原表 UUID，未自動合併。"
        elif len(candidates) > 1:
            match_status = "AMBIGUOUS_MULTIPLE_CANDIDATES"
            match_method = "KEY_CATEGORY_MULTIPLE"
            staging_uuid = item["source_uid"]
            original_uuid = ""
            type_label, name, county, town = item["type"], item["name"], item["county"], item["town"]
            original_lon = original_lat = None
            original_row = ""
            note = "同縣市、鄉鎮、地名與類別仍有多筆候選，未自動合併。"
        else:
            match_status = "NO_MATCH"
            match_method = "NO_KEY_CATEGORY_MATCH"
            staging_uuid = item["source_uid"]
            original_uuid = ""
            type_label, name, county, town = item["type"], item["name"], item["county"], item["town"]
            original_lon = original_lat = None
            original_row = ""
            note = "找不到同縣市、鄉鎮、地名與類別候選，使用新清冊 UID。"

        new_lat, new_lon, coordinate_status = item["new_coordinates"]
        if new_lat is not None and new_lon is not None:
            proposed_lon, proposed_lat = new_lon, new_lat
        elif match_status == "UNIQUE_MATCH" and original_lat is not None and original_lon is not None:
            proposed_lon, proposed_lat = original_lon, original_lat
            coordinate_status = "NEW_INVALID_PRESERVED_ORIGINAL"
        else:
            proposed_lon = proposed_lat = None
            coordinate_status = "NEW_INVALID_LEFT_BLANK"

        values = [
            match_status,
            match_method,
            staging_uuid,
            original_uuid,
            item["source_uid"],
            item["place_id"],
            item["code"],
            type_label,
            name,
            county,
            town,
            item["village"],
            item["location"],
            item["history"],
            item["standard_code"],
            item["data_source"],
            original_lon,
            original_lat,
            proposed_lon,
            proposed_lat,
            coordinate_status,
            item["candidate_count"],
            target_group_size,
            original_row,
            note,
            batch_name,
        ]
        for index, value in enumerate(values):
            if isinstance(value, str) and len(value) > 49000:
                values[index] = value[:49000]
                overlong_cells += 1
        output_rows.append(values)
        status_counts[match_status] += 1
        coordinate_counts[coordinate_status] += 1

    summary = {
        "live_rows": len(live_rows),
        "new_rows": len(new_rows),
        "staging_rows": len(output_rows),
        "status_counts": dict(status_counts),
        "coordinate_counts": dict(coordinate_counts),
        "duplicate_live_uuids": sum(1 for count in live_uuid_counts.values() if count > 1),
        "overlong_cells_truncated": overlong_cells,
        "headers": HEADERS,
    }
    return output_rows, summary


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--live-snapshot", type=Path, required=True)
    parser.add_argument("--new-inventory", type=Path, required=True)
    parser.add_argument("--output-jsonl", type=Path, required=True)
    parser.add_argument("--output-compact-jsonl", type=Path)
    parser.add_argument("--output-compact-xlsx", type=Path)
    parser.add_argument("--compact-chunk-dir", type=Path)
    parser.add_argument("--compact-chunk-rows", type=int, default=1000)
    parser.add_argument("--summary-json", type=Path, required=True)
    parser.add_argument("--batch-name", default="115.1.12_地名後臺清冊_合併")
    args = parser.parse_args()

    live_rows = load_live_snapshot(args.live_snapshot)
    new_rows = load_new_inventory(args.new_inventory)
    output_rows, summary = build_staging(live_rows, new_rows, args.batch_name)

    args.output_jsonl.parent.mkdir(parents=True, exist_ok=True)
    with args.output_jsonl.open("w", encoding="utf-8") as destination:
        destination.write(json.dumps(HEADERS, ensure_ascii=False) + "\n")
        for row in output_rows:
            destination.write(json.dumps(row, ensure_ascii=False, default=str) + "\n")

    if args.output_compact_jsonl:
        indexes = {header: index for index, header in enumerate(HEADERS)}
        compact_rows = []
        with args.output_compact_jsonl.open("w", encoding="utf-8") as destination:
            destination.write(json.dumps(COMPACT_HEADERS, ensure_ascii=False) + "\n")
            for row in output_rows:
                match_info = {
                    "status": row[indexes["MatchStatus"]],
                    "method": row[indexes["MatchMethod"]],
                    "source_uid": row[indexes["NewSourceUID"]],
                    "source_place_id": row[indexes["NewPlaceId"]],
                    "category_code": row[indexes["NewCategoryCode"]],
                    "type": row[indexes["Type"]],
                    "place_name": row[indexes["PlaceName"]],
                    "county": row[indexes["County"]],
                    "town": row[indexes["Town"]],
                    "candidate_count": row[indexes["CandidateCount"]],
                    "new_rows_for_target": row[indexes["NewRowsForTarget"]],
                    "original_row": row[indexes["OriginalRow"]],
                    "note": row[indexes["MatchNote"]],
                    "batch": row[indexes["ImportBatch"]],
                }
                coordinates = {
                    "original_longitude": row[indexes["OriginalLongitude"]],
                    "original_latitude": row[indexes["OriginalLatitude"]],
                    "proposed_longitude": row[indexes["ProposedLongitude"]],
                    "proposed_latitude": row[indexes["ProposedLatitude"]],
                    "status": row[indexes["CoordinateStatus"]],
                }
                compact_row = [
                    json.dumps(match_info, ensure_ascii=False, separators=(",", ":")),
                    row[indexes["StagingUUID"]],
                    row[indexes["Village"]],
                    row[indexes["LocationDescription"]],
                    row[indexes["HistoryDescription"]],
                    row[indexes["StandardPlaceCode"]],
                    row[indexes["DataSource"]],
                    json.dumps(coordinates, ensure_ascii=False, separators=(",", ":")),
                ]
                destination.write(json.dumps(compact_row, ensure_ascii=False, default=str) + "\n")
                compact_rows.append(compact_row)

        if args.compact_chunk_dir:
            args.compact_chunk_dir.mkdir(parents=True, exist_ok=True)
            rows_per_chunk = max(args.compact_chunk_rows, 1)
            all_compact_rows = [COMPACT_HEADERS, *compact_rows]
            for start in range(0, len(all_compact_rows), rows_per_chunk):
                chunk_number = start // rows_per_chunk + 1
                chunk_path = args.compact_chunk_dir / f"chunk-{chunk_number:04d}.json"
                chunk_path.write_text(
                    json.dumps(all_compact_rows[start:start + rows_per_chunk], ensure_ascii=False, default=str),
                    encoding="utf-8",
                )

        if args.output_compact_xlsx:
            args.output_compact_xlsx.parent.mkdir(parents=True, exist_ok=True)
            workbook = Workbook(write_only=True)
            sheet = workbook.create_sheet("staging")
            sheet.append(COMPACT_HEADERS)
            for compact_row in compact_rows:
                sheet.append(compact_row)
            workbook.save(args.output_compact_xlsx)
    args.summary_json.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
